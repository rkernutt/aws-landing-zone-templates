#!/usr/bin/env python3
"""Build the three workbook templates: Accounts, Subnet Plan, CIS Controls Tracker."""
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "/sessions/amazing-sharp-bell/mnt/AWS Design Docs/Templates"

NAVY = "1F3864"
ACCENT = "2E75B6"
LIGHT = "D5E8F0"
ALT = "F5F8FC"
RED = "C00000"
GREEN = "385723"

def header_font():
    return Font(name="Arial", color="FFFFFF", bold=True, size=11)

def body_font():
    return Font(name="Arial", size=10)

def placeholder_font():
    return Font(name="Arial", color=RED, italic=True, bold=True, size=10)

def title_font():
    return Font(name="Arial", color=NAVY, bold=True, size=16)

def section_font():
    return Font(name="Arial", color=NAVY, bold=True, size=12)

def small_font():
    return Font(name="Arial", color="595959", italic=True, size=9)

def header_fill():
    return PatternFill("solid", start_color=NAVY)

def alt_fill():
    return PatternFill("solid", start_color=ALT)

def light_fill():
    return PatternFill("solid", start_color=LIGHT)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def write_headers(ws, row, headers, widths=None):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = header_font()
        c.fill = header_fill()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[row].height = 28
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

def write_row(ws, row, values, fill=None, placeholder_cols=None):
    placeholder_cols = placeholder_cols or set()
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        is_placeholder = (isinstance(v, str) and ("{{" in v)) or (i in placeholder_cols)
        c.font = placeholder_font() if is_placeholder else body_font()
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = thin_border()
        if fill:
            c.fill = fill

def add_title(ws, title, subtitle, ncols):
    ws["A1"] = title
    ws["A1"].font = title_font()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.row_dimensions[1].height = 28
    ws["A2"] = subtitle
    ws["A2"].font = small_font()
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)

# ---------- Workbook 1: Accounts ----------
def build_accounts():
    wb = Workbook()
    # Cover
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "AWS Landing Zone — Accounts & Tag Policies"
    cover["A1"].font = title_font()
    cover.merge_cells("A1:F1")
    cover["A3"] = "Customer:"
    cover["B3"] = "{{CUSTOMER_NAME}}"
    cover["A4"] = "Partner:"
    cover["B4"] = "{{PARTNER_NAME}}"
    cover["A5"] = "Document version:"
    cover["B5"] = "{{DOC_VERSION}}"
    cover["A6"] = "Date:"
    cover["B6"] = "{{DOC_DATE}}"
    cover["A7"] = "Classification:"
    cover["B7"] = "{{DOC_CLASSIFICATION}}"
    for r in range(3, 8):
        cover.cell(row=r, column=1).font = section_font()
        cell = cover.cell(row=r, column=2)
        cell.font = placeholder_font()
    cover["A9"] = "How to use this workbook"
    cover["A9"].font = section_font()
    instructions = [
        "1. Replace all {{PLACEHOLDERS}} (red italic text) with values for the target customer.",
        "2. The 'Accounts' sheet lists every AWS account created in the Landing Zone.",
        "3. The 'OUs' sheet documents the Organisational Unit structure.",
        "4. The 'Tag Policies' sheet captures the mandatory tagging schema enforced by Organizations tag policies.",
        "5. The 'SCP Attachment' sheet shows which SCP / RCP set applies to each OU.",
        "6. Keep account IDs in TEXT format (Excel will otherwise strip leading zeros).",
        "7. Email addresses follow the pattern aws+<purpose>@{{CUSTOMER_DOMAIN}}.",
    ]
    for i, t in enumerate(instructions, start=10):
        c = cover.cell(row=i, column=1, value=t)
        c.font = body_font()
        cover.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
    cover.column_dimensions["A"].width = 22
    cover.column_dimensions["B"].width = 60
    cover.sheet_view.showGridLines = False

    # Accounts sheet
    accts = wb.create_sheet("Accounts")
    add_title(accts, "Accounts", "All AWS accounts created by the Landing Zone. Foundation accounts first, then workload accounts grouped by OU.", 8)
    headers = ["Account Purpose", "Alias", "Root Email", "Account ID", "OU", "Owner Team", "Environment", "Tags"]
    widths = [32, 26, 38, 16, 26, 22, 16, 32]
    write_headers(accts, 4, headers, widths)
    # Freeze + filter
    accts.freeze_panes = "A5"
    # Foundation rows
    foundation = [
        ("Management (Org root)",        "{{CUSTOMER_CODE}}-management",     "aws+management@{{CUSTOMER_DOMAIN}}",  "{{MANAGEMENT_ACCOUNT_ID}}", "Root",            "Cloud Platform", "n/a",        "BusinessUnit=Platform; Environment=Management"),
        ("Log Archive",                  "{{CUSTOMER_CODE}}-log-archive",    "aws+logs@{{CUSTOMER_DOMAIN}}",        "{{LOG_ARCHIVE_ACCOUNT_ID}}", "Security",        "Security",       "n/a",        "BusinessUnit=Platform; Environment=Security"),
        ("Audit (Security Tooling)",     "{{CUSTOMER_CODE}}-audit",          "aws+audit@{{CUSTOMER_DOMAIN}}",       "{{AUDIT_ACCOUNT_ID}}",       "Security",        "Security",       "n/a",        "BusinessUnit=Platform; Environment=Security"),
        ("Network",                      "{{CUSTOMER_CODE}}-network",        "aws+network@{{CUSTOMER_DOMAIN}}",     "{{NETWORK_ACCOUNT_ID}}",     "Infrastructure",  "Network",        "n/a",        "BusinessUnit=Platform; Environment=Infrastructure"),
        ("Shared Services",              "{{CUSTOMER_CODE}}-shared-services","aws+shared@{{CUSTOMER_DOMAIN}}",      "{{SHARED_SERVICES_ACCOUNT_ID}}", "Infrastructure", "Cloud Platform", "n/a",        "BusinessUnit=Platform; Environment=Infrastructure"),
        ("Backup (optional)",            "{{CUSTOMER_CODE}}-backup",         "aws+backup@{{CUSTOMER_DOMAIN}}",      "TBD",                        "Infrastructure",  "Cloud Platform", "n/a",        "BusinessUnit=Platform; Environment=Infrastructure"),
    ]
    r = 5
    for row in foundation:
        write_row(accts, r, row, fill=light_fill() if r % 2 == 1 else None)
        r += 1
    # Example workload accounts
    workload_examples = [
        ("Example BU Production",      "{{CUSTOMER_CODE}}-<bu>-prod",     "aws+<bu>-prod@{{CUSTOMER_DOMAIN}}",     "TBD", "Workloads/Production",    "<BU> Engineering", "Production",     "BusinessUnit=<BU>; Environment=Production"),
        ("Example BU Pre-Production",  "{{CUSTOMER_CODE}}-<bu>-preprod",  "aws+<bu>-preprod@{{CUSTOMER_DOMAIN}}",  "TBD", "Workloads/NonProduction", "<BU> Engineering", "PreProduction",  "BusinessUnit=<BU>; Environment=PreProduction"),
        ("Example BU UAT",             "{{CUSTOMER_CODE}}-<bu>-uat",      "aws+<bu>-uat@{{CUSTOMER_DOMAIN}}",      "TBD", "Workloads/NonProduction", "<BU> Engineering", "UAT",            "BusinessUnit=<BU>; Environment=UAT"),
        ("Example BU SIT",             "{{CUSTOMER_CODE}}-<bu>-sit",      "aws+<bu>-sit@{{CUSTOMER_DOMAIN}}",      "TBD", "Workloads/NonProduction", "<BU> Engineering", "SIT",            "BusinessUnit=<BU>; Environment=SIT"),
        ("Example BU Test",            "{{CUSTOMER_CODE}}-<bu>-test",     "aws+<bu>-test@{{CUSTOMER_DOMAIN}}",     "TBD", "Workloads/NonProduction", "<BU> Engineering", "Test",           "BusinessUnit=<BU>; Environment=Test"),
        ("Example BU Development",     "{{CUSTOMER_CODE}}-<bu>-dev",      "aws+<bu>-dev@{{CUSTOMER_DOMAIN}}",      "TBD", "Workloads/NonProduction", "<BU> Engineering", "Development",    "BusinessUnit=<BU>; Environment=Development"),
        ("Example Sandbox",            "{{CUSTOMER_CODE}}-sandbox-<user>","aws+sandbox-<user>@{{CUSTOMER_DOMAIN}}","TBD", "Sandbox",                 "Cloud Platform",   "Sandbox",        "BusinessUnit=Sandbox; Environment=Sandbox"),
    ]
    for row in workload_examples:
        write_row(accts, r, row, fill=alt_fill() if r % 2 == 1 else None)
        r += 1
    # Validation for OU and Environment columns
    ou_list = '"Root,Security,Infrastructure,Workloads/Production,Workloads/NonProduction,Sandbox,PolicyStaging,Suspended"'
    env_list = '"Production,PreProduction,UAT,SIT,Test,Development,Sandbox,n/a"'
    dv_ou = DataValidation(type="list", formula1=ou_list, allow_blank=True)
    dv_env = DataValidation(type="list", formula1=env_list, allow_blank=True)
    accts.add_data_validation(dv_ou); accts.add_data_validation(dv_env)
    dv_ou.add(f"E5:E200"); dv_env.add(f"G5:G200")
    # Account ID column as text
    for row in accts.iter_rows(min_row=5, min_col=4, max_col=4, max_row=200):
        for c in row:
            c.number_format = "@"
    accts.auto_filter.ref = f"A4:H{r-1}"

    # OUs sheet
    ous = wb.create_sheet("OUs")
    add_title(ous, "Organisational Units", "OU structure deployed via Control Tower / Organizations.", 4)
    write_headers(ous, 4, ["OU", "Parent", "Purpose", "SCP / RCP set"], [28, 22, 50, 30])
    ous.freeze_panes = "A5"
    ou_rows = [
        ("Root", "—", "Top-level container.", "Root SCP/RCP baseline"),
        ("Security", "Root", "Houses Log Archive and Audit accounts.", "Security OU SCPs (most restrictive)"),
        ("Infrastructure", "Root", "Network and Shared Services accounts.", "Infrastructure OU SCPs"),
        ("Workloads", "Root", "Parent container for workload OUs.", "Inherited from Root"),
        ("Workloads/Production", "Workloads", "Production workload accounts grouped by BU.", "Production OU SCPs"),
        ("Workloads/NonProduction", "Workloads", "Non-production workload accounts grouped by BU.", "Non-Production OU SCPs"),
        ("Sandbox", "Root", "Throwaway dev / PoC accounts. Tight budget guardrails.", "Sandbox SCPs (tightest egress)"),
        ("PolicyStaging", "Root", "OU for testing new SCP / RCP changes before promotion.", "Staging SCPs"),
        ("Suspended", "Root", "Decommissioned accounts pending closure.", "DenyAll SCP"),
    ]
    rr = 5
    for row in ou_rows:
        write_row(ous, rr, row, fill=alt_fill() if rr % 2 == 1 else None)
        rr += 1

    # Tag Policies sheet
    tags = wb.create_sheet("Tag Policies")
    add_title(tags, "Tag Policies", "Mandatory tag keys enforced by AWS Organizations tag policies.", 5)
    write_headers(tags, 4, ["Tag Key", "Description", "Allowed Values", "Enforced On", "Required"], [24, 40, 40, 22, 12])
    tags.freeze_panes = "A5"
    tag_rows = [
        ("BusinessUnit",      "Owning business unit", "{{BUSINESS_UNITS}}", "All resources", "Yes"),
        ("Environment",       "Lifecycle environment", "Production,PreProduction,UAT,SIT,Test,Development,Sandbox", "All resources", "Yes"),
        ("CostCentre",        "Internal cost code", "<customer cost code list>", "All resources", "Yes"),
        ("DataClassification","Sensitivity classification", "Public,Internal,Confidential,Restricted", "Data-bearing resources", "Yes"),
        ("Owner",             "Email / DL of resource owner", "<email>", "All resources", "Yes"),
        ("Project",           "Project / programme name", "<free text>", "All resources", "No"),
        ("Compliance",        "Regulatory tag", "PCI,GDPR,SOX,FCA,None", "Selected resources", "No"),
        ("BackupPolicy",      "AWS Backup tag", "daily-7d,daily-35d,monthly-12m,yearly-7y,none", "Backup-eligible", "Conditional"),
    ]
    rr = 5
    for row in tag_rows:
        write_row(tags, rr, row, fill=alt_fill() if rr % 2 == 1 else None)
        rr += 1

    # SCP Attachment sheet
    scp = wb.create_sheet("SCP Attachment")
    add_title(scp, "SCP / RCP Attachment Plan", "Which policy bundles attach to which OU.", 4)
    write_headers(scp, 4, ["OU", "Policy Type", "Policy Name", "Notes"], [28, 18, 36, 40])
    scp.freeze_panes = "A5"
    scp_rows = [
        ("Root", "SCP", "DenyRootUser", "Block root credentials."),
        ("Root", "SCP", "RegionRestriction", "Allow only {{PRIMARY_REGION}} + {{SECONDARY_REGION}}."),
        ("Root", "SCP", "DenyLeaveOrganization", "Prevent account from leaving Org."),
        ("Root", "SCP", "DenyDisableSecurityServices", "Block disabling Security Hub / GuardDuty / Config / CloudTrail."),
        ("Root", "RCP", "EnforceTLSEverywhere", "Require aws:SecureTransport=true on S3/SQS/SNS."),
        ("Root", "RCP", "DenyPublicResourcePolicies", "Block public access on resource policies."),
        ("Security", "SCP", "SecurityOUGuardrails", "Tighter restrictions on security accounts."),
        ("Infrastructure", "SCP", "InfraOUGuardrails", "Network change controls."),
        ("Workloads/Production", "SCP", "ProdGuardrails", "Restricts destructive actions."),
        ("Workloads/NonProduction", "SCP", "NonProdGuardrails", "Looser; allows dev experimentation."),
        ("Sandbox", "SCP", "SandboxGuardrails", "Aggressive: no IAM user creation, capped instance sizes."),
    ]
    rr = 5
    for row in scp_rows:
        write_row(scp, rr, row, fill=alt_fill() if rr % 2 == 1 else None)
        rr += 1

    out = os.path.join(OUT, "06-Accounts-Template.xlsx")
    wb.save(out)
    print("Wrote", out)

# ---------- Workbook 2: Subnet Plan ----------
def build_subnets():
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "AWS Landing Zone — Subnet Plan"
    cover["A1"].font = title_font()
    cover.merge_cells("A1:G1")
    meta = [
        ("Customer:", "{{CUSTOMER_NAME}}"),
        ("Partner:",  "{{PARTNER_NAME}}"),
        ("Region:",   "{{PRIMARY_REGION}}"),
        ("Supernet:", "{{LZ_SUPERNET}}"),
        ("Date:",     "{{DOC_DATE}}"),
    ]
    for i, (k, v) in enumerate(meta, start=3):
        cover.cell(row=i, column=1, value=k).font = section_font()
        c = cover.cell(row=i, column=2, value=v); c.font = placeholder_font()
    cover["A9"] = "How to use this workbook"
    cover["A9"].font = section_font()
    notes = [
        "1. The 'Region Allocation' sheet records the /16 (or larger) supernet per region.",
        "2. 'OU Allocation' shows the /19 slices per OU / Business Unit.",
        "3. 'BU-<name>' sheets show per-environment, per-tier subnets (Web/App/DB × AZ-a/AZ-b/AZ-c).",
        "4. 'Network Services' shows the Network account inspection VPC subnets.",
        "5. 'Shared Services' shows the Shared Services VPC.",
        "6. 'Sandbox' shows the Sandbox CIDR pool.",
        "7. Update the supernet on Cover and OU Allocation; the BU sheets show the recommended /21 split per environment.",
        "8. Hosts column shows usable hosts (= 2^(32-mask) - 5; AWS reserves the first 4 and last 1).",
    ]
    for i, t in enumerate(notes, start=10):
        cover.cell(row=i, column=1, value=t).font = body_font()
        cover.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)
    cover.column_dimensions["A"].width = 22
    cover.column_dimensions["B"].width = 50
    cover.sheet_view.showGridLines = False

    # Region allocation
    reg = wb.create_sheet("Region Allocation")
    add_title(reg, "Region Allocation", "Per-region /16 allocations from the LZ supernet.", 5)
    write_headers(reg, 4, ["Region", "Region Name", "CIDR", "Status", "Notes"], [16, 28, 22, 14, 36])
    rows = [
        ("{{PRIMARY_REGION}}", "{{PRIMARY_REGION_NAME}}", "{{LZ_REGION_CIDR}}", "Active", "Day-one primary region."),
        ("{{SECONDARY_REGION}}", "(future / DR)",          "TBD",                  "Planned", "Reserve a /16 in the supernet for DR or expansion."),
    ]
    rr = 5
    for row in rows:
        write_row(reg, rr, row, fill=alt_fill() if rr % 2 == 1 else None)
        rr += 1

    # OU allocation
    ou = wb.create_sheet("OU Allocation")
    add_title(ou, "OU / Business Unit Allocation", "Per-OU /19 slices from the regional /16. Example shown using 172.26.0.0/16.", 5)
    write_headers(ou, 4, ["CIDR", "Mask", "Hosts (usable)", "Allocation", "Notes"], [22, 10, 16, 36, 26])
    ou_rows = [
        ("172.26.0.0/19",   "/19", 8190, "Network, Shared Services, Sandbox", "Foundation pool"),
        ("172.26.32.0/19",  "/19", 8190, "Business Unit #1 — example: Asset Finance", ""),
        ("172.26.64.0/19",  "/19", 8190, "Business Unit #2 — example: Motor Finance", ""),
        ("172.26.96.0/19",  "/19", 8190, "Business Unit #3 — example: Premium Finance", ""),
        ("172.26.128.0/19", "/19", 8190, "Business Unit #4 — example: Invoice Finance", ""),
        ("172.26.160.0/19", "/19", 8190, "Business Unit #5 — example: Treasury", ""),
        ("172.26.192.0/19", "/19", 8190, "Business Unit #6 — example: Property", ""),
        ("172.26.224.0/19", "/19", 8190, "Business Unit #7 — example: Central Services", ""),
    ]
    rr = 5
    for row in ou_rows:
        write_row(ou, rr, row, fill=alt_fill() if rr % 2 == 1 else None)
        rr += 1

    # Per-BU environment template
    bu_envs = [
        ("Production",    "/21", "{{CUSTOMER_CODE}}-<bu>-prod"),
        ("PreProduction", "/21", "{{CUSTOMER_CODE}}-<bu>-preprod"),
        ("UAT",           "/22", "{{CUSTOMER_CODE}}-<bu>-uat"),
        ("SIT",           "/22", "{{CUSTOMER_CODE}}-<bu>-sit"),
        ("Test",          "/22", "{{CUSTOMER_CODE}}-<bu>-test"),
        ("Development",   "/22", "{{CUSTOMER_CODE}}-<bu>-dev"),
    ]
    tier_layouts = {
        "/21": [
            ("Web A", "/24", 254, "eu-west-2a"), ("Web B", "/24", 254, "eu-west-2b"),
            ("App A", "/24", 254, "eu-west-2a"), ("App B", "/24", 254, "eu-west-2b"),
            ("DB A",  "/24", 254, "eu-west-2a"), ("DB B",  "/24", 254, "eu-west-2b"),
            ("TGW A","/28",   11, "eu-west-2a"), ("TGW B","/28",   11, "eu-west-2b"),
        ],
        "/22": [
            ("Web A", "/25", 126, "eu-west-2a"), ("Web B", "/25", 126, "eu-west-2b"),
            ("App A", "/25", 126, "eu-west-2a"), ("App B", "/25", 126, "eu-west-2b"),
            ("DB A",  "/25", 126, "eu-west-2a"), ("DB B",  "/25", 126, "eu-west-2b"),
            ("TGW A","/28",   11, "eu-west-2a"), ("TGW B","/28",   11, "eu-west-2b"),
        ],
    }
    ws = wb.create_sheet("BU-Template")
    add_title(ws, "Business Unit Subnet Template", "Copy this sheet per business unit; rename and adjust CIDRs.", 8)
    write_headers(ws, 4, ["Environment", "VPC Name", "VPC CIDR", "Subnet Name", "Subnet CIDR", "Tier", "AZ", "Hosts"], [18, 26, 18, 32, 22, 14, 16, 10])
    ws.freeze_panes = "A5"
    rr = 5
    for env, mask, vpc in bu_envs:
        for tier, sub_mask, hosts, az in tier_layouts[mask]:
            subnet_name = f"{vpc}-{tier.lower().replace(' ', '-')}"
            write_row(ws, rr, [env, vpc, f"172.26.x.x{mask}", subnet_name, f"172.26.x.x{sub_mask}", tier.split()[0], az, hosts],
                      fill=alt_fill() if rr % 2 == 1 else None)
            rr += 1

    # Network Services
    ns = wb.create_sheet("Network Services")
    add_title(ns, "Network Services VPCs", "Inspection VPCs in the Network account.", 7)
    write_headers(ns, 4, ["VPC", "Purpose", "VPC CIDR", "Subnet", "Subnet CIDR", "AZ", "Notes"], [28, 24, 16, 26, 18, 16, 30])
    ns_rows = [
        ("{{CUSTOMER_CODE}}-inspection-egress",   "Outbound traffic inspection",   "172.26.1.0/24", "egress-fw-a",  "172.26.1.0/27",   "eu-west-2a", "Firewall ENI A"),
        ("{{CUSTOMER_CODE}}-inspection-egress",   "Outbound traffic inspection",   "172.26.1.0/24", "egress-fw-b",  "172.26.1.32/27",  "eu-west-2b", "Firewall ENI B"),
        ("{{CUSTOMER_CODE}}-inspection-egress",   "Outbound traffic inspection",   "172.26.1.0/24", "egress-tgw-a", "172.26.1.64/27",  "eu-west-2a", "TGW attach A"),
        ("{{CUSTOMER_CODE}}-inspection-egress",   "Outbound traffic inspection",   "172.26.1.0/24", "egress-tgw-b", "172.26.1.96/27",  "eu-west-2b", "TGW attach B"),
        ("{{CUSTOMER_CODE}}-inspection-egress",   "Outbound traffic inspection",   "172.26.1.0/24", "egress-pub-a", "172.26.1.128/27", "eu-west-2a", "NAT GW + IGW"),
        ("{{CUSTOMER_CODE}}-inspection-egress",   "Outbound traffic inspection",   "172.26.1.0/24", "egress-pub-b", "172.26.1.160/27", "eu-west-2b", "NAT GW + IGW"),
        ("{{CUSTOMER_CODE}}-inspection-ingress",  "Inbound traffic inspection",    "172.26.0.0/24", "ingress-pub-a","172.26.0.0/27",   "eu-west-2a", "Public ALB / IGW"),
        ("{{CUSTOMER_CODE}}-inspection-ingress",  "Inbound traffic inspection",    "172.26.0.0/24", "ingress-pub-b","172.26.0.32/27",  "eu-west-2b", "Public ALB / IGW"),
        ("{{CUSTOMER_CODE}}-inspection-eastwest", "East/West inspection",          "172.26.2.0/24", "ew-fw-a",      "172.26.2.0/27",   "eu-west-2a", "Firewall ENI A"),
        ("{{CUSTOMER_CODE}}-inspection-eastwest", "East/West inspection",          "172.26.2.0/24", "ew-fw-b",      "172.26.2.32/27",  "eu-west-2b", "Firewall ENI B"),
    ]
    rr = 5
    for row in ns_rows:
        write_row(ns, rr, row, fill=alt_fill() if rr % 2 == 1 else None)
        rr += 1

    # Shared Services
    ss = wb.create_sheet("Shared Services")
    add_title(ss, "Shared Services VPC", "Cross-account shared services workloads.", 6)
    write_headers(ss, 4, ["VPC", "Subnet", "CIDR", "Tier", "AZ", "Purpose"], [28, 28, 18, 14, 16, 28])
    ss_rows = [
        ("{{CUSTOMER_CODE}}-shared-services", "shared-web-a", "172.26.4.0/25",   "Web",  "eu-west-2a", "Public-facing entry points"),
        ("{{CUSTOMER_CODE}}-shared-services", "shared-web-b", "172.26.4.128/25", "Web",  "eu-west-2b", "Public-facing entry points"),
        ("{{CUSTOMER_CODE}}-shared-services", "shared-app-a", "172.26.5.0/25",   "App",  "eu-west-2a", "Internal applications"),
        ("{{CUSTOMER_CODE}}-shared-services", "shared-app-b", "172.26.5.128/25", "App",  "eu-west-2b", "Internal applications"),
        ("{{CUSTOMER_CODE}}-shared-services", "shared-db-a",  "172.26.6.0/25",   "DB",   "eu-west-2a", "Shared data stores"),
        ("{{CUSTOMER_CODE}}-shared-services", "shared-db-b",  "172.26.6.128/25", "DB",   "eu-west-2b", "Shared data stores"),
    ]
    rr = 5
    for row in ss_rows:
        write_row(ss, rr, row, fill=alt_fill() if rr % 2 == 1 else None)
        rr += 1

    # Sandbox
    sb = wb.create_sheet("Sandbox")
    add_title(sb, "Sandbox CIDR Pool", "Pool of /22 ranges that can be allocated to individual sandbox accounts.", 4)
    write_headers(sb, 4, ["CIDR", "Hosts (usable)", "Status", "Notes"], [22, 18, 16, 40])
    sb_rows = [
        ("172.26.8.0/22",  1022, "Available", "Assign to first sandbox account"),
        ("172.26.12.0/22", 1022, "Available", ""),
        ("172.26.16.0/22", 1022, "Available", ""),
        ("172.26.20.0/22", 1022, "Available", ""),
        ("172.26.24.0/22", 1022, "Available", ""),
        ("172.26.28.0/22", 1022, "Available", ""),
    ]
    rr = 5
    for row in sb_rows:
        write_row(sb, rr, row, fill=alt_fill() if rr % 2 == 1 else None)
        rr += 1

    out = os.path.join(OUT, "07-Subnet-Plan-Template.xlsx")
    wb.save(out)
    print("Wrote", out)

# ---------- Workbook 3: CIS Controls Tracker ----------
def build_cis():
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "AWS Security Controls Tracker"
    cover["A1"].font = title_font()
    cover.merge_cells("A1:G1")
    cover["A2"] = "CIS AWS Foundations Benchmark v3 + FSBP + customer compliance frameworks"
    cover["A2"].font = small_font()
    cover.merge_cells("A2:G2")
    meta = [
        ("Customer:",  "{{CUSTOMER_NAME}}"),
        ("Partner:",   "{{PARTNER_NAME}}"),
        ("Frameworks:", "{{COMPLIANCE_FRAMEWORKS}}"),
        ("Date:",      "{{DOC_DATE}}"),
    ]
    for i, (k, v) in enumerate(meta, start=4):
        cover.cell(row=i, column=1, value=k).font = section_font()
        cover.cell(row=i, column=2, value=v).font = placeholder_font()
    cover["A9"] = "How to use this workbook"
    cover["A9"].font = section_font()
    notes = [
        "1. 'CIS v3 Controls' lists every control with implementation responsibility, status, evidence and notes.",
        "2. 'Summary' aggregates status counts and gives a compliance posture view.",
        "3. Status options: Compliant, Partial, Non-Compliant, Not Applicable, In Progress.",
        "4. Responsibility options: Customer, Partner, Shared, AWS.",
        "5. Update 'Status' as controls are validated; the Summary recalculates automatically.",
        "6. Add additional sheets for other frameworks (NIST CSF 2.0, ISO 27001, PCI DSS 4.0).",
    ]
    for i, t in enumerate(notes, start=10):
        cover.cell(row=i, column=1, value=t).font = body_font()
        cover.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)
    cover.column_dimensions["A"].width = 22
    cover.column_dimensions["B"].width = 56
    cover.sheet_view.showGridLines = False

    # CIS Controls
    ws = wb.create_sheet("CIS v3 Controls")
    add_title(ws, "CIS AWS Foundations Benchmark v3 — Control Status Tracker", "One row per control. Update Status, Evidence and Owner per delivery.", 8)
    headers = ["Control #", "Title", "Level", "Section", "Responsibility", "Status", "Evidence / Source", "Notes"]
    widths = [12, 56, 8, 24, 18, 18, 30, 30]
    write_headers(ws, 4, headers, widths)
    ws.freeze_panes = "A5"

    # Subset / representative CIS v3 control list (Identity, Storage, Logging, Monitoring, Networking)
    controls = [
        # Identity & Access Management
        ("1.1", "Maintain current contact details", 1, "Identity & Access Management"),
        ("1.2", "Ensure security contact information is registered", 1, "Identity & Access Management"),
        ("1.3", "Ensure security questions are registered in the AWS account", 1, "Identity & Access Management"),
        ("1.4", "Ensure no 'root' user account access key exists", 1, "Identity & Access Management"),
        ("1.5", "Ensure MFA is enabled for the 'root' user account", 1, "Identity & Access Management"),
        ("1.6", "Ensure hardware MFA is enabled for the 'root' user account", 2, "Identity & Access Management"),
        ("1.7", "Eliminate use of the 'root' user for administrative and daily tasks", 1, "Identity & Access Management"),
        ("1.8", "Ensure IAM password policy requires minimum length of 14 or greater", 1, "Identity & Access Management"),
        ("1.9", "Ensure IAM password policy prevents password reuse", 1, "Identity & Access Management"),
        ("1.10", "Ensure MFA is enabled for all IAM users with a console password", 1, "Identity & Access Management"),
        ("1.11", "Do not create access keys during initial setup for IAM users with a console password", 1, "Identity & Access Management"),
        ("1.12", "Ensure credentials unused for 45 days or greater are disabled", 1, "Identity & Access Management"),
        ("1.13", "Ensure there is only one active access key for any single IAM user", 1, "Identity & Access Management"),
        ("1.14", "Ensure access keys are rotated every 90 days or less", 1, "Identity & Access Management"),
        ("1.15", "Ensure IAM users receive permissions only through groups", 1, "Identity & Access Management"),
        ("1.16", "Ensure IAM policies that allow full ':' administrative privileges are not attached", 1, "Identity & Access Management"),
        ("1.17", "Ensure a support role has been created to manage incidents with AWS Support", 1, "Identity & Access Management"),
        ("1.18", "Ensure IAM instance roles are used for AWS resource access from instances", 2, "Identity & Access Management"),
        ("1.19", "Ensure that all the expired SSL/TLS certificates stored in AWS IAM are removed", 1, "Identity & Access Management"),
        ("1.20", "Ensure IAM Access Analyzer is enabled for all regions", 1, "Identity & Access Management"),
        ("1.21", "Ensure IAM users are managed centrally via identity federation or AWS IAM Identity Center", 2, "Identity & Access Management"),
        ("1.22", "Ensure access to AWSCloudShellFullAccess is restricted", 1, "Identity & Access Management"),
        # Storage
        ("2.1.1", "Ensure S3 buckets employ encryption-at-rest", 1, "Storage"),
        ("2.1.2", "Ensure S3 Bucket Policy is set to deny HTTP requests", 1, "Storage"),
        ("2.1.3", "Ensure MFA Delete is enabled on S3 buckets", 2, "Storage"),
        ("2.1.4", "Ensure all data in S3 has been discovered, classified and secured", 2, "Storage"),
        ("2.1.5", "Ensure S3 Block Public Access setting is enabled at the account level", 1, "Storage"),
        ("2.2.1", "Ensure EBS volume encryption is enabled in all regions", 1, "Storage"),
        ("2.3.1", "Ensure encryption-at-rest is enabled for RDS instances", 1, "Storage"),
        ("2.3.2", "Ensure auto minor version upgrade is enabled for RDS instances", 1, "Storage"),
        ("2.3.3", "Ensure RDS instances are not publicly accessible", 1, "Storage"),
        ("2.4.1", "Ensure encryption is enabled for EFS file systems", 1, "Storage"),
        # Logging
        ("3.1", "Ensure CloudTrail is enabled in all regions", 1, "Logging"),
        ("3.2", "Ensure CloudTrail log file validation is enabled", 2, "Logging"),
        ("3.3", "Ensure the S3 bucket used to store CloudTrail logs is not publicly accessible", 1, "Logging"),
        ("3.4", "Ensure CloudTrail trails are integrated with CloudWatch Logs", 1, "Logging"),
        ("3.5", "Ensure AWS Config is enabled in all regions", 1, "Logging"),
        ("3.6", "Ensure S3 bucket access logging is enabled on the CloudTrail S3 bucket", 1, "Logging"),
        ("3.7", "Ensure CloudTrail logs are encrypted at rest using KMS CMKs", 2, "Logging"),
        ("3.8", "Ensure rotation for customer-created symmetric CMKs is enabled", 1, "Logging"),
        ("3.9", "Ensure VPC Flow Logs is enabled in all VPCs", 1, "Logging"),
        ("3.10", "Ensure object-level logging for write events is enabled for S3 bucket", 2, "Logging"),
        ("3.11", "Ensure object-level logging for read events is enabled for S3 bucket", 2, "Logging"),
        # Monitoring
        ("4.1", "Ensure unauthorized API calls are monitored", 1, "Monitoring"),
        ("4.2", "Ensure management console sign-in without MFA is monitored", 1, "Monitoring"),
        ("4.3", "Ensure usage of 'root' account is monitored", 1, "Monitoring"),
        ("4.4", "Ensure IAM policy changes are monitored", 1, "Monitoring"),
        ("4.5", "Ensure CloudTrail configuration changes are monitored", 1, "Monitoring"),
        ("4.6", "Ensure AWS Management Console authentication failures are monitored", 2, "Monitoring"),
        ("4.7", "Ensure disabling/scheduled deletion of customer-created CMKs is monitored", 2, "Monitoring"),
        ("4.8", "Ensure S3 bucket policy changes are monitored", 1, "Monitoring"),
        ("4.9", "Ensure AWS Config configuration changes are monitored", 2, "Monitoring"),
        ("4.10", "Ensure security group changes are monitored", 2, "Monitoring"),
        ("4.11", "Ensure NACL changes are monitored", 2, "Monitoring"),
        ("4.12", "Ensure changes to network gateways are monitored", 1, "Monitoring"),
        ("4.13", "Ensure route table changes are monitored", 1, "Monitoring"),
        ("4.14", "Ensure VPC changes are monitored", 1, "Monitoring"),
        ("4.15", "Ensure AWS Organizations changes are monitored", 1, "Monitoring"),
        ("4.16", "Ensure AWS Security Hub is enabled", 2, "Monitoring"),
        # Networking
        ("5.1", "Ensure no NACL allows ingress from 0.0.0.0/0 to remote server administration ports", 1, "Networking"),
        ("5.2", "Ensure no security groups allow ingress from 0.0.0.0/0 to remote server administration ports", 1, "Networking"),
        ("5.3", "Ensure no security groups allow ingress from ::/0 to remote server administration ports", 1, "Networking"),
        ("5.4", "Ensure the default security group of every VPC restricts all traffic", 2, "Networking"),
        ("5.5", "Ensure routing tables for VPC peering are 'least access'", 2, "Networking"),
        ("5.6", "Ensure that EC2 Metadata Service v2 (IMDSv2) is enabled and required", 1, "Networking"),
    ]
    rr = 5
    for ctrl in controls:
        write_row(ws, rr, [ctrl[0], ctrl[1], ctrl[2], ctrl[3], "Shared", "In Progress", "", ""], fill=alt_fill() if rr % 2 == 1 else None)
        rr += 1

    # Data validation
    status_list = '"Compliant,Partial,Non-Compliant,Not Applicable,In Progress"'
    resp_list = '"Customer,Partner,Shared,AWS"'
    dv_status = DataValidation(type="list", formula1=status_list, allow_blank=True)
    dv_resp = DataValidation(type="list", formula1=resp_list, allow_blank=True)
    ws.add_data_validation(dv_status); ws.add_data_validation(dv_resp)
    dv_status.add(f"F5:F{rr-1}"); dv_resp.add(f"E5:E{rr-1}")
    ws.auto_filter.ref = f"A4:H{rr-1}"

    # Conditional formatting via patterns — simpler: leave for end user, rely on status text + filter.
    # Summary sheet
    summ = wb.create_sheet("Summary")
    add_title(summ, "Summary", "Auto-aggregated status counts across all controls.", 3)
    write_headers(summ, 4, ["Status", "Count", "% of Total"], [22, 12, 14])
    statuses = ["Compliant", "Partial", "Non-Compliant", "Not Applicable", "In Progress"]
    total_cell = f"'CIS v3 Controls'!F5:F{rr-1}"
    for i, st in enumerate(statuses, start=5):
        summ.cell(row=i, column=1, value=st).font = body_font()
        summ.cell(row=i, column=1).border = thin_border()
        cnt = summ.cell(row=i, column=2, value=f'=COUNTIF({total_cell},"{st}")')
        cnt.font = body_font(); cnt.border = thin_border()
        pct = summ.cell(row=i, column=3, value=f'=IFERROR(B{i}/SUM($B$5:$B$9),0)')
        pct.font = body_font(); pct.border = thin_border()
        pct.number_format = "0.0%"
    # Total row
    summ.cell(row=10, column=1, value="Total").font = section_font()
    summ.cell(row=10, column=2, value=f"=SUM(B5:B9)").font = section_font()
    summ.cell(row=10, column=3, value=f"=SUM(C5:C9)").font = section_font()
    summ.cell(row=10, column=3).number_format = "0.0%"
    for col in range(1, 4):
        summ.cell(row=10, column=col).border = thin_border()
        summ.cell(row=10, column=col).fill = light_fill()
    # Section breakdown
    summ["A12"] = "Section Breakdown"
    summ["A12"].font = section_font()
    write_headers(summ, 13, ["Section", "Compliant", "Partial", "Non-Compliant", "Not Applicable", "In Progress", "Total"], [28, 12, 12, 14, 14, 14, 10])
    sections = ["Identity & Access Management", "Storage", "Logging", "Monitoring", "Networking"]
    section_col = f"'CIS v3 Controls'!D5:D{rr-1}"
    status_col = f"'CIS v3 Controls'!F5:F{rr-1}"
    for i, sec in enumerate(sections, start=14):
        summ.cell(row=i, column=1, value=sec).font = body_font(); summ.cell(row=i, column=1).border = thin_border()
        for j, st in enumerate(statuses, start=2):
            c = summ.cell(row=i, column=j, value=f'=COUNTIFS({section_col},"{sec}",{status_col},"{st}")')
            c.font = body_font(); c.border = thin_border()
        t = summ.cell(row=i, column=7, value=f"=SUM(B{i}:F{i})")
        t.font = body_font(); t.border = thin_border()

    out = os.path.join(OUT, "08-CIS-Controls-Tracker-Template.xlsx")
    wb.save(out)
    print("Wrote", out)

if __name__ == "__main__":
    build_accounts()
    build_subnets()
    build_cis()
